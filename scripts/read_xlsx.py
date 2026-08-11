#!/usr/bin/env python3
"""读 xlsx 内容到文件，避免 Windows console 编码问题"""
import sys
import io
import openpyxl

# 强制 UTF-8 输出
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = sys.stdout

def dump_xlsx(path, out_path, max_rows=200):
    if path.endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        lines = []
        for sheet in wb.sheets():
            lines.append(f"=== Sheet: {sheet.name} (rows={sheet.nrows}, cols={sheet.ncols}) ===")
            for r in range(min(sheet.nrows, max_rows)):
                row = sheet.row_values(r)
                if any(str(c).strip() for c in row):
                    lines.append("  | " + " | ".join(str(c) for c in row))
            if sheet.nrows > max_rows:
                lines.append(f"  ... (truncated at row {max_rows} of {sheet.nrows})")
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===")
            for r, row in enumerate(ws.iter_rows(values_only=True), 1):
                if r > max_rows:
                    lines.append(f"  ... (truncated at row {max_rows})")
                    break
                if any(c is not None and str(c).strip() != '' for c in row):
                    cells = [str(c) if c is not None else '' for c in row]
                    lines.append("  | " + " | ".join(cells))
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))
    print(f"Written: {out_path}")

if __name__ == "__main__":
    dump_xlsx(r"E:\junzhijia\公司资料\公司资料\军之甲产品与业务概况.xlsx",
              r"C:\Users\Administrator\.mavis\workspace\box-beeaa-site\data\jzcj-overview.txt",
              max_rows=100)
    dump_xlsx(r"E:\junzhijia\公司资料\公司资料\最新伟立产品规格表.xls",
              r"C:\Users\Administrator\.mavis\workspace\box-beeaa-site\data\wl-products.txt",
              max_rows=300)
